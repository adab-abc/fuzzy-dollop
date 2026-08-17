# -*- coding: utf-8 -*-
"""
Created on Fri Aug  7 11:47:32 2026

@author: adabreo
"""
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment

HEADER_FONT = Font(bold=True,size=14)
SUBHEADER_FONT = Font(bold=True,size=11)
PAIR_FONT = Font(bold=True)
PASS_FILL = PatternFill(fill_type="solid",start_color="C6EFCE",end_color="C6EFCE")
GRAY_FILL = PatternFill(fill_type="solid",start_color="808080",end_color="808080")

def reportStandardCurvePlot(analysisData):

    curveFitData = analysisData["IMG"]["curveFitData"]

    fig, axes = plt.subplots(
        2,
        2,
        figsize=(8, 6)
    )

    axes = axes.flatten()

    for axis, pairName in zip(axes, curveFitData.index):

        xValues = []
        yValues = []

        for concentrationLabel in curveFitData.columns:

            yValue = curveFitData.loc[pairName, concentrationLabel]

            if pd.notna(yValue):

                concentration = float(
                    str(concentrationLabel)
                    .replace(" ng/ml", "")
                    .replace(" mg/ml", "")
                )

                xValues.append(concentration)

                yValues.append(yValue)

        axis.plot(
            xValues,
            yValues,
            marker="o",
            linewidth=1
        )

        axis.set_xscale("log")
        axis.set_yscale("log")

        axis.set_title(
            pairName,
            fontsize=8
        )

        axis.grid(True)

    plt.tight_layout()

    plt.savefig(
        "standard_curve.png",
        dpi=250,
        bbox_inches="tight"
    )

    plt.close()
def reportModelComparison(analysisData):
    print("\nMODEL COMPARISON\n")
    for pairName, comparison in analysisData["IMG"]["modelComparison"].items():
        print(pairName)
        print("4PL:", comparison["fourPLRSquared"])
        print("5PL:", comparison["fivePLRSquared"])
        print("Delta:", comparison["deltaRSquared"])
        print()

def reportInterpolation(sampleName, analysisData):
    print(f"\n{sampleName} RESULTS\n")
    print(analysisData[sampleName]["interpolatedData"])
    
def reportLinearity(sampleName, analysisData):
    print(f"\n{sampleName} LINEARITY MATRICES\n")
    for pairName, matrix in analysisData[sampleName]["linearityMatrices"].items():
        print(pairName)
        print(matrix)
        print()

def generateReport(analysisData):
    reportModelComparison(analysisData)
    reportStandardCurvePlot(analysisData)
    for sampleName in ["AP", "Sample1", "Sample2"]:
        reportInterpolation(sampleName, analysisData)
    for sampleName in ["AP", "Sample1", "Sample2"]:
        reportLinearity(sampleName, analysisData)
    print("\nREPORT GENERATED")
    

def writeDataFrame(reportSheet, dataframe, startRow, startColumn, applyLinearityFormatting=False):
    for columnOffset, columnName in enumerate(dataframe.columns, start=1):
        reportSheet.cell(
            row=startRow,
            column=startColumn + columnOffset,
            value=str(columnName)
        )
    for rowOffset, rowName in enumerate(dataframe.index, start=1):
        reportSheet.cell(
            row=startRow + rowOffset,
            column=startColumn,
            value=str(rowName)
        )
        for columnOffset, columnName in enumerate(dataframe.columns, start=1):

            value = dataframe.loc[rowName, columnName]
        
            cell = reportSheet.cell(
                row=startRow + rowOffset,
                column=startColumn + columnOffset,
                value=value
            )
        
            cell.alignment = Alignment(horizontal="center")
        
            if applyLinearityFormatting:
        
                if pd.isna(value):
        
                    cell.fill = GRAY_FILL
        
                elif 70 <= value <= 130:
        
                    cell.fill = PASS_FILL
                    
    return startRow + len(dataframe.index) + 1

def writeSampleSection(
    reportSheet,
    sampleName,
    analysisData,
    currentRow
):

    cell = reportSheet.cell(currentRow, 1, sampleName)
    cell.font = HEADER_FONT
    currentRow += 2

    cell = reportSheet.cell(
        currentRow,
        1,
        "INTERPOLATED DATA (pg/mL)"
    )
    cell.font = SUBHEADER_FONT

    cell = reportSheet.cell(
        currentRow,
        14,
        "DATA CORRECTED BY DILUTION FACTOR (pg/mL)"
    )
    cell.font = SUBHEADER_FONT

    currentRow += 2

    interpolatedData = analysisData[sampleName]["interpolatedData"].copy()
    correctedData = analysisData[sampleName]["correctedData"].copy()

    writeDataFrame(
        reportSheet,
        interpolatedData,
        currentRow,
        1
    )

    writeDataFrame(
        reportSheet,
        correctedData,
        currentRow,
        14
    )

    tableHeight = len(interpolatedData.index) + 3

    currentRow += tableHeight + 1

    cell = reportSheet.cell(
        currentRow,
        1,
        "LINEARITY MATRICES (%)"
    )
    cell.font = SUBHEADER_FONT

    currentRow += 1

    matrices = list(
        analysisData[sampleName]["linearityMatrices"].items()
    )

    for i in range(0, len(matrices), 2):

        leftPair, leftMatrix = matrices[i]

        rightPair = None
        rightMatrix = None

        if i + 1 < len(matrices):
            rightPair, rightMatrix = matrices[i + 1]

        cell = reportSheet.cell(currentRow, 1, leftPair)
        cell.font = PAIR_FONT

        if rightPair is not None:
            cell = reportSheet.cell(currentRow, 14, rightPair)
            cell.font = PAIR_FONT

        currentRow += 1

        writeDataFrame(
            reportSheet,
            leftMatrix,
            currentRow,
            1,
            applyLinearityFormatting=True
        )

        if rightMatrix is not None:

            writeDataFrame(
                reportSheet,
                rightMatrix,
                currentRow,
                14,
                applyLinearityFormatting=True
            )

        matrixHeight = len(leftMatrix.index) + 3

        currentRow += matrixHeight + 1

    return currentRow

def generateExcelReport(
    workbookPath,
    worksheetName,
    analysisData
):
    if workbookPath.lower().endswith(".xlsm"):

        workbook = load_workbook(
            workbookPath,
            keep_vba=True
    )

    else:
    
        workbook = load_workbook(workbookPath)
    reportSheetName = f"{worksheetName}_Report"
    if reportSheetName in workbook.sheetnames:
        del workbook[reportSheetName]
    workbook.create_sheet(reportSheetName)
    reportSheet = workbook[reportSheetName]
    
    # ==========================
    # MODEL COMPARISON
    # ==========================
    reportSheet["A1"] = "MODEL COMPARISON"
    reportSheet["A1"].font = HEADER_FONT
    reportSheet["A3"] = "Pair"
    reportSheet["B3"] = "4PL R²"
    reportSheet["C3"] = "5PL R²"
    reportSheet["D3"] = "Delta"
    currentRow = 4
    for pairName, comparison in analysisData["IMG"]["modelComparison"].items():
        reportSheet.cell(currentRow, 1, pairName)
        reportSheet.cell(currentRow, 2, comparison["fourPLRSquared"])
        reportSheet.cell(currentRow, 3, comparison["fivePLRSquared"])
        reportSheet.cell(currentRow, 4, comparison["deltaRSquared"])
        currentRow += 1

    # ==========================
    # STANDARD CURVE PLOT
    # ==========================
    reportStandardCurvePlot(analysisData)
    reportSheet["A10"] = "STANDARD CURVE PLOT"
    reportSheet["A10"].font = HEADER_FONT
    plotImage = Image("standard_curve.png")
    
    plotImage.width = 500
    plotImage.height = 350
    
    reportSheet.add_image(
        plotImage,
        "A12"
    )

    # ==========================
    # SAMPLE SECTIONS
    # ==========================
    currentRow = 29
    
    for sampleName in ["AP", "Sample1", "Sample2"]:
    
        currentRow = writeSampleSection(
            reportSheet,
            sampleName,
            analysisData,
            currentRow
        )
    
        currentRow += 1

    # Pair-name columns

    reportSheet.column_dimensions["A"].width = 40
    reportSheet.column_dimensions["N"].width = 40
    
    # Left-side data columns
    
    for column in ["B", "C", "D", "E", "F", "G", "H", "I", "J"]:
    
        reportSheet.column_dimensions[column].width = 12
    
    # Right-side data columns
    
    for column in ["O", "P", "Q", "R", "S", "T", "U", "V", "W"]:
    
        reportSheet.column_dimensions[column].width = 12
    #Save Report Worksheet
    workbook.save(workbookPath)
    print(f"\nCreated worksheet: {reportSheetName}")
